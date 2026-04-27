"""Excel 导出:把 output_items_dir 下的 JSON 合并成业务可读的 Excel。

读 config.yaml 的 output_items_dir 和 excel_output_dir,生成
result_YYYYMMDD_HHMM.xlsx,含三个 sheet:
- 主表:一份文档一行,含元信息/硬约束/质量分/摘要/状态/复核列
- 问题清单:评估发现的具体问题,一条一行
- 失败列表:抽取失败的文档

启动方式同 runner.py:
    cd tasks/01_std_eval && python excel_export.py
    python tasks/01_std_eval/excel_export.py tasks/01_std_eval

注意:此脚本不调用 LLM,只做本地数据合并。修改 Excel 格式不需要重跑 runner。
"""
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from shared.config_loader import load_config

sys.stdout.reconfigure(line_buffering=True)


# -------------------- 颜色和样式常量 --------------------

COLOR_HEADER = "B4C7E7"   # 表头蓝灰
COLOR_HIT    = "F4B6B6"   # 硬约束命中红
COLOR_FAILED = "FFB6B6"   # 失败行红
COLOR_REVIEW = "FFF2CC"   # 人工复核列浅黄

# 质量分 1-5 的渐变色
SCORE_COLORS = {
    5: "C6EFCE",  # 深绿
    4: "E2EFDA",  # 浅绿
    3: "FFEB9C",  # 黄
    2: "FCD5B4",  # 橙
    1: "F4B6B6",  # 红
}

HARD_CONSTRAINTS = ["内部矛盾", "附件图片缺失", "引用失效", "占位符残留", "歧义无法执行"]

# 主表列定义:(列标题, 列宽)
MAIN_COLUMNS: list[tuple[str, int]] = [
    # 基础信息
    ("文件名",     30),
    ("标准名称",   30),
    ("一句话概括", 40),
    ("文档摘要",   60),
    # 元信息
    ("标准类型",        10),
    ("是否监管要求",    10),
    ("监管置信度",      10),
    ("可检查性主导类型", 12),
    # 硬约束(5 项)
    ("内部矛盾",     20),
    ("附件图片缺失", 20),
    ("引用失效",     20),
    ("占位符残留",   20),
    ("歧义无法执行", 20),
    # 质量分(3 项)
    ("表述清晰度",   8),
    ("结构完整性",   8),
    ("可操作性",     8),
    # 状态
    ("任务状态",     10),
    # 人工复核(最后两列)
    ("复核结论",     12),
    ("复核备注",     30),
]
REVIEW_COLUMN_COUNT = 2  # 末尾几列是复核列(浅黄底,失败时不覆盖)


# -------------------- 通用样式工具 --------------------

def _border() -> Border:
    thin = Side(border_style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _set_header_row(ws: Worksheet, columns: list[tuple[str, int]]) -> None:
    """渲染表头(第一行)"""
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    font = Font(bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, (name, _w) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _set_column_widths(ws: Worksheet, columns: list[tuple[str, int]]) -> None:
    for col_idx, (_n, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# -------------------- 数据加载 --------------------

def load_items(items_dir: Path) -> list[dict]:
    """读 items 目录下所有 JSON,返回字典列表。

    损坏文件(读不出 JSON)会被跳过并打印警告,不中断流程。
    """
    if not items_dir.is_dir():
        raise FileNotFoundError(f"items 目录不存在: {items_dir}")

    records = []
    bad_files = []
    for path in sorted(items_dir.glob("*.json")):
        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            bad_files.append((path.name, str(e)))
            continue
        if isinstance(data, dict):
            records.append(data)

    if bad_files:
        print(f"  ⚠ 跳过 {len(bad_files)} 份损坏文件:")
        for name, err in bad_files:
            print(f"    - {name}: {err}")

    return records


# -------------------- 单元格填充工具 --------------------

def _fill_hard_constraint(cell, hit_info: dict) -> None:
    """硬约束单元格:命中显示 "是\\n证据: xxx"(红底),未命中显示 "否"。"""
    if not isinstance(hit_info, dict):
        cell.value = ""
        return
    if hit_info.get("命中") is True:
        evidence = hit_info.get("证据", "")
        cell.value = f"是\n证据: {evidence}" if evidence else "是"
        cell.fill = PatternFill("solid", fgColor=COLOR_HIT)
    else:
        cell.value = "否"


def _fill_score(cell, score_info: dict, *, is_principle_doc: bool, is_operability: bool) -> None:
    """质量分单元格:按分数着色。原则性文档的可操作性显示 N/A。"""
    if is_operability and is_principle_doc:
        cell.value = "N/A"
        return
    if not isinstance(score_info, dict):
        cell.value = ""
        return
    score = score_info.get("建议分")
    if score is None:
        cell.value = "N/A"
        return
    cell.value = score
    color = SCORE_COLORS.get(score)
    if color:
        cell.fill = PatternFill("solid", fgColor=color)


# -------------------- 主表 sheet --------------------

def _get_status(record: dict) -> str:
    if record.get("_status") == "failed":
        return "失败"
    return "成功"


def write_main_sheet(ws: Worksheet, records: list[dict]) -> None:
    ws.title = "主表"
    _set_header_row(ws, MAIN_COLUMNS)

    border = _border()
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    review_fill = PatternFill("solid", fgColor=COLOR_REVIEW)
    failed_fill = PatternFill("solid", fgColor=COLOR_FAILED)

    # 按 source_file 排序保证稳定输出
    records_sorted = sorted(records, key=lambda r: r.get("source_file", ""))

    for row_idx, rec in enumerate(records_sorted, start=2):
        status = _get_status(rec)
        is_failed = (status == "失败")

        # 失败记录:基础信息空着,只显示文件名和状态
        meta = rec.get("元信息", {}) if not is_failed else {}
        reg = meta.get("是否监管要求", {}) if isinstance(meta.get("是否监管要求"), dict) else {}
        std_type = meta.get("标准类型", "")
        is_principle = (std_type == "原则性")
        hard = rec.get("硬约束", {}) if not is_failed else {}
        scores = rec.get("质量分", {}) if not is_failed else {}

        col = 1
        def put(value, *, alignment=align_left):
            nonlocal col
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = alignment
            cell.border = border
            col += 1
            return cell

        # 基础信息
        put(rec.get("source_file", ""))
        put(rec.get("标准名称", ""))
        put(rec.get("一句话概括", "") if not is_failed else "")
        put(rec.get("文档摘要", "") if not is_failed else "")

        # 元信息
        put(std_type)
        put(reg.get("建议", ""))
        put(reg.get("置信度", ""))
        put(meta.get("可检查性主导类型", ""))

        # 硬约束 5 列
        for hc_name in HARD_CONSTRAINTS:
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = align_left
            cell.border = border
            if not is_failed:
                _fill_hard_constraint(cell, hard.get(hc_name, {}))
            col += 1

        # 质量分 3 列
        for qs_name in ["表述清晰度", "结构完整性", "可操作性"]:
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = align_center
            cell.border = border
            if not is_failed:
                _fill_score(
                    cell,
                    scores.get(qs_name, {}),
                    is_principle_doc=is_principle,
                    is_operability=(qs_name == "可操作性"),
                )
            col += 1

        # 状态
        put(status, alignment=align_center)

        # 复核列(浅黄底,留空)
        for _ in range(REVIEW_COLUMN_COUNT):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = align_left
            cell.border = border
            cell.fill = review_fill
            col += 1

        # 失败行整行覆盖红色,但保留末尾 REVIEW_COLUMN_COUNT 列的浅黄
        if is_failed:
            non_review_count = len(MAIN_COLUMNS) - REVIEW_COLUMN_COUNT
            for c in range(1, non_review_count + 1):
                ws.cell(row=row_idx, column=c).fill = failed_fill

    _set_column_widths(ws, MAIN_COLUMNS)
    ws.freeze_panes = "B2"  # 冻结首行 + 文件名列


# -------------------- 问题清单 sheet --------------------

PROBLEM_COLUMNS: list[tuple[str, int]] = [
    ("文件名",   30),
    ("问题序号", 8),
    ("章节",     15),
    ("描述",     40),
    ("证据",     40),
    ("复核结论", 12),
    ("复核备注", 30),
]


def write_problems_sheet(ws: Worksheet, records: list[dict]) -> None:
    ws.title = "问题清单"
    _set_header_row(ws, PROBLEM_COLUMNS)

    border = _border()
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    review_fill = PatternFill("solid", fgColor=COLOR_REVIEW)

    row = 2
    records_sorted = sorted(records, key=lambda r: r.get("source_file", ""))

    for rec in records_sorted:
        if rec.get("_status") == "failed":
            continue
        problems = rec.get("问题清单", [])
        if not isinstance(problems, list):
            continue
        for idx, p in enumerate(problems, start=1):
            if not isinstance(p, dict):
                continue
            values = [
                rec.get("source_file", ""),
                idx,
                p.get("章节", ""),
                p.get("描述", ""),
                p.get("证据", ""),
                "",  # 复核结论
                "",  # 复核备注
            ]
            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.alignment = align_left
                cell.border = border
                # 末尾两列复核浅黄
                if col_idx > len(PROBLEM_COLUMNS) - REVIEW_COLUMN_COUNT:
                    cell.fill = review_fill
            row += 1

    _set_column_widths(ws, PROBLEM_COLUMNS)
    ws.freeze_panes = "B2"


# -------------------- 失败列表 sheet --------------------

FAIL_COLUMNS: list[tuple[str, int]] = [
    ("文件名",   30),
    ("错误信息", 60),
    ("时间戳",   20),
]


def write_failures_sheet(ws: Worksheet, records: list[dict]) -> None:
    ws.title = "失败列表"
    _set_header_row(ws, FAIL_COLUMNS)

    border = _border()
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row = 2
    failed = [r for r in records if r.get("_status") == "failed"]
    failed.sort(key=lambda r: r.get("source_file", ""))

    for rec in failed:
        values = [
            rec.get("source_file", ""),
            rec.get("_error", ""),
            rec.get("_timestamp", ""),
        ]
        for col_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.alignment = align_left
            cell.border = border
        row += 1

    _set_column_widths(ws, FAIL_COLUMNS)


# -------------------- 主入口 --------------------

def _resolve_task_dir() -> Path:
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).resolve()
    return Path(__file__).resolve().parent


def main():
    task_dir = _resolve_task_dir()
    config = load_config(task_dir / "config.yaml")

    items_dir_str = config.get("output_items_dir", "").strip()
    excel_dir_str = config.get("excel_output_dir", "").strip()
    if not items_dir_str:
        print("config.yaml 的 output_items_dir 未配置")
        sys.exit(1)
    if not excel_dir_str:
        print("config.yaml 的 excel_output_dir 未配置")
        sys.exit(1)
    items_dir = Path(items_dir_str)
    excel_dir = Path(excel_dir_str)

    print(f"读取 items: {items_dir}")
    records = load_items(items_dir)
    if not records:
        print("没有可用的记录,退出")
        sys.exit(0)

    succeeded = sum(1 for r in records if r.get("_status") != "failed")
    failed = len(records) - succeeded
    print(f"加载 {len(records)} 条记录(成功 {succeeded},失败 {failed})")

    wb = Workbook()
    write_main_sheet(wb.active, records)
    write_problems_sheet(wb.create_sheet(), records)
    write_failures_sheet(wb.create_sheet(), records)

    excel_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = excel_dir / f"result_{timestamp}.xlsx"
    wb.save(str(output_path))
    print(f"已生成: {output_path}")


if __name__ == "__main__":
    main()
