"""
Excel 生成脚本:合并评估结果和摘要结果,生成业务人工复核用的 Excel

输入:
- EVAL_MERGED: 评估 merged.json(来自 eval_items + template_merge)
- SUMMARY_MERGED: 摘要 merged.json(来自 summary_items + template_merge)

输出:
- result_YYYYMMDD_HHMM.xlsx,含三个 sheet:
  * 主表:一份文档一行,含元信息/硬约束/质量分/摘要/复核列
  * 问题清单:评估发现的具体问题,一条一行
  * 失败列表:评估或摘要失败的文档

合并逻辑:按 source_file 字段 join。缺一方的文档也会出现在主表里。

依赖:
    openpyxl
"""
import os
import sys
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(line_buffering=True)

# ============= 配置区 =============
EVAL_MERGED    = r"C:\std_gov\eval_merged.json"
SUMMARY_MERGED = r"C:\std_gov\summary_merged.json"
OUTPUT_DIR     = r"C:\std_gov"
# ==================================

# 颜色定义
COLOR_HEADER   = "B4C7E7"    # 表头蓝灰
COLOR_GROUP    = "D9E1F2"    # 分组副表头浅蓝
COLOR_HIT      = "F4B6B6"    # 硬约束命中红
COLOR_FAILED   = "FFB6B6"    # 失败行红
COLOR_REVIEW   = "FFF2CC"    # 人工复核列浅黄

# 质量分颜色映射
SCORE_COLORS = {
    5: "C6EFCE",  # 深绿
    4: "E2EFDA",  # 浅绿
    3: "FFEB9C",  # 黄
    2: "FCD5B4",  # 橙
    1: "F4B6B6",  # 红
}

# 硬约束项目顺序
HARD_CONSTRAINTS = ["内部矛盾", "附件图片缺失", "引用失效", "占位符残留", "歧义无法执行"]

# 主表列定义(列标题, 列宽)
MAIN_COLUMNS = [
    # 基础信息
    ("文件名",     30),
    ("标准名称",   30),
    ("一句话概括", 40),
    ("文档摘要",   60),
    # 元信息
    ("标准类型",        10),
    ("是否监管要求",    10),
    ("监管置信度",      10),
    ("可检查性主导类型", 12),
    # 硬约束(5 项)
    ("内部矛盾",     20),
    ("附件图片缺失", 20),
    ("引用失效",     20),
    ("占位符残留",   20),
    ("歧义无法执行", 20),
    # 质量分(3 项)
    ("表述清晰度",   8),
    ("结构完整性",   8),
    ("可操作性",     8),
    # 状态
    ("评估状态",     10),
    ("摘要状态",     10),
    # 人工复核
    ("复核结论",     12),
    ("复核备注",     30),
]


# -------------------- 加载数据 --------------------

def load_merged(path: str) -> list:
    if not os.path.isfile(path):
        print(f"警告: 未找到 {path}")
        return []
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        print(f"警告: {path} 顶层不是数组")
        return []
    return data


def build_lookup(records: list) -> dict:
    """按 source_file 建索引,便于 join"""
    return {r.get("source_file", ""): r for r in records if isinstance(r, dict)}


# -------------------- 单元格填充工具 --------------------

def fill_hard_constraint(cell, hit_info: dict):
    """硬约束单元格:命中显示 "是\n证据: xxx",未命中显示 "否" """
    if not isinstance(hit_info, dict):
        cell.value = ""
        return
    hit = hit_info.get("命中")
    evidence = hit_info.get("证据", "")
    if hit is True:
        cell.value = f"是\n证据: {evidence}" if evidence else "是"
        cell.fill = PatternFill("solid", fgColor=COLOR_HIT)
    else:
        cell.value = "否"


def fill_score(cell, score_info: dict, is_principle_doc: bool, is_operability: bool):
    """质量分单元格:按分数着色。原则性文档的可操作性显示 N/A"""
    if is_operability and is_principle_doc:
        cell.value = "N/A"
        return
    if not isinstance(score_info, dict):
        cell.value = ""
        return
    score = score_info.get("建议分")
    if score is None:
        cell.value = "N/A"
        return
    cell.value = score
    color = SCORE_COLORS.get(score)
    if color:
        cell.fill = PatternFill("solid", fgColor=color)


# -------------------- 写主表 --------------------

def get_status(record: dict) -> str:
    """获取一份记录的状态"""
    if not record:
        return "未跑"
    if record.get("_status") == "failed":
        return "失败"
    return "成功"


def write_main_sheet(ws, all_files: list, eval_lookup: dict, summary_lookup: dict):
    # 表头
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, _width) in enumerate(MAIN_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 数据行
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    review_fill = PatternFill("solid", fgColor=COLOR_REVIEW)
    failed_fill = PatternFill("solid", fgColor=COLOR_FAILED)

    for row_idx, fname in enumerate(all_files, start=2):
        eval_rec = eval_lookup.get(fname, {})
        sum_rec = summary_lookup.get(fname, {})

        eval_status = get_status(eval_rec)
        sum_status = get_status(sum_rec)
        row_failed = (eval_status == "失败" or sum_status == "失败")

        # 基础字段
        standard_name = (eval_rec.get("标准名称")
                         or sum_rec.get("标准名称")
                         or "")
        one_liner = sum_rec.get("一句话概括", "") if sum_status == "成功" else ""
        summary = sum_rec.get("文档摘要", "") if sum_status == "成功" else ""

        # 元信息(评估里)
        meta = eval_rec.get("元信息", {}) if eval_status == "成功" else {}
        std_type = meta.get("标准类型", "")
        reg = meta.get("是否监管要求", {}) if isinstance(meta.get("是否监管要求"), dict) else {}
        reg_answer = reg.get("建议", "")
        reg_conf = reg.get("置信度", "")
        checkability = meta.get("可检查性主导类型", "")

        # 硬约束
        hard = eval_rec.get("硬约束", {}) if eval_status == "成功" else {}

        # 质量分
        scores = eval_rec.get("质量分", {}) if eval_status == "成功" else {}
        is_principle = (std_type == "原则性")

        # 按列写入
        col = 1

        def put(v):
            nonlocal col
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.alignment = left_wrap
            cell.border = border
            col += 1
            return cell

        put(fname)
        put(standard_name)
        put(one_liner)
        put(summary)
        put(std_type)
        put(reg_answer)
        put(reg_conf)
        put(checkability)

        # 硬约束 5 列
        for hc_name in HARD_CONSTRAINTS:
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = left_wrap
            cell.border = border
            fill_hard_constraint(cell, hard.get(hc_name, {}))
            col += 1

        # 质量分 3 列
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = center
        cell.border = border
        fill_score(cell, scores.get("表述清晰度", {}), is_principle, is_operability=False)
        col += 1

        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = center
        cell.border = border
        fill_score(cell, scores.get("结构完整性", {}), is_principle, is_operability=False)
        col += 1

        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = center
        cell.border = border
        fill_score(cell, scores.get("可操作性", {}), is_principle, is_operability=True)
        col += 1

        # 状态列
        put(eval_status)
        put(sum_status)

        # 复核列(留白,浅黄底)
        for _ in range(2):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = left_wrap
            cell.border = border
            cell.fill = review_fill
            col += 1

        # 失败行全行覆盖红色(覆盖之前的硬约束/分数着色)
        if row_failed:
            for c in range(1, len(MAIN_COLUMNS) + 1):
                # 保留复核列的浅黄色
                if c >= len(MAIN_COLUMNS) - 1:
                    continue
                ws.cell(row=row_idx, column=c).fill = failed_fill

    # 列宽
    for col_idx, (_col_name, width) in enumerate(MAIN_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行 + 文件名列
    ws.freeze_panes = "B2"


# -------------------- 写问题清单 sheet --------------------

PROBLEM_COLUMNS = [
    ("文件名",   30),
    ("问题序号", 8),
    ("章节",     15),
    ("描述",     40),
    ("证据",     40),
    ("复核结论", 12),
    ("复核备注", 30),
]


def write_problems_sheet(ws, eval_lookup: dict):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, _w) in enumerate(PROBLEM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    review_fill = PatternFill("solid", fgColor=COLOR_REVIEW)

    row = 2
    for fname in sorted(eval_lookup.keys()):
        rec = eval_lookup[fname]
        if not isinstance(rec, dict) or rec.get("_status") == "failed":
            continue
        problems = rec.get("问题清单", [])
        if not isinstance(problems, list):
            continue
        for idx, p in enumerate(problems, start=1):
            if not isinstance(p, dict):
                continue
            values = [
                fname,
                idx,
                p.get("章节", ""),
                p.get("描述", ""),
                p.get("证据", ""),
                "",  # 复核结论
                "",  # 复核备注
            ]
            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.alignment = left_wrap
                cell.border = border
                # 复核列浅黄
                if col_idx >= len(PROBLEM_COLUMNS) - 1:
                    cell.fill = review_fill
            row += 1

    for col_idx, (_col_name, width) in enumerate(PROBLEM_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "B2"


# -------------------- 写失败列表 sheet --------------------

FAIL_COLUMNS = [
    ("文件名",   30),
    ("失败来源", 12),
    ("错误信息", 60),
    ("时间戳",   20),
]


def write_failures_sheet(ws, eval_lookup: dict, summary_lookup: dict):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, _w) in enumerate(FAIL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row = 2
    for source_name, lookup in [("评估", eval_lookup), ("摘要", summary_lookup)]:
        for fname in sorted(lookup.keys()):
            rec = lookup[fname]
            if not isinstance(rec, dict) or rec.get("_status") != "failed":
                continue
            values = [
                fname,
                source_name,
                rec.get("_error", ""),
                rec.get("_timestamp", ""),
            ]
            for col_idx, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=v)
                cell.alignment = left_wrap
                cell.border = border
            row += 1

    for col_idx, (_col_name, width) in enumerate(FAIL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# -------------------- 主流程 --------------------

def main():
    eval_records = load_merged(EVAL_MERGED)
    summary_records = load_merged(SUMMARY_MERGED)

    if not eval_records and not summary_records:
        print("两份 merged.json 都没数据,退出")
        return

    eval_lookup = build_lookup(eval_records)
    summary_lookup = build_lookup(summary_records)

    # 主表的行数 = 两边并集,按文件名排序
    all_files = sorted(set(eval_lookup.keys()) | set(summary_lookup.keys()))

    print(f"评估记录: {len(eval_records)} 条")
    print(f"摘要记录: {len(summary_records)} 条")
    print(f"主表总行数(并集): {len(all_files)}")

    wb = Workbook()

    ws_main = wb.active
    ws_main.title = "主表"
    write_main_sheet(ws_main, all_files, eval_lookup, summary_lookup)

    ws_problems = wb.create_sheet("问题清单")
    write_problems_sheet(ws_problems, eval_lookup)

    ws_fails = wb.create_sheet("失败列表")
    write_failures_sheet(ws_fails, eval_lookup, summary_lookup)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"result_{timestamp}.xlsx")
    wb.save(output_path)

    print(f"\n已生成: {output_path}")


if __name__ == "__main__":
    main()
